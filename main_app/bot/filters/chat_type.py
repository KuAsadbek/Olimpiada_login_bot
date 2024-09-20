from aiogram.types import Message,ContentType
from aiogram.filters import Filter
from aiogram.filters import BaseFilter

import openpyxl
from openpyxl.styles import Alignment


import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

# Функция для стилизации ячеек с заголовками
def style_header_cells(sheet, columns):
    header_font = Font(bold=True, size=12, color="FFFFFF")  # Белый жирный шрифт
    fill = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")  # Синий фон
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Тонкие границы
    
    for col in columns:
        for cell in sheet[col][:1]:  # Только первая строка (заголовки)
            cell.font = header_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Центровка

# Функция для добавления границ и выравнивания
def style_body_cells(sheet, columns):
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in columns:
        sheet.column_dimensions[col].width = 20  # Задаем ширину столбцов
        for cell in sheet[col][1:]:  # Все строки, кроме первой (заголовков)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")  # Выравнивание по левому краю

# Функция для раскрашивания ячеек в колонке 'G' в зависимости от значения
def style_column_g(sheet):
    green_fill = PatternFill(fill_type="solid", start_color="00FF00", end_color="00FF00")  # Зеленый цвет
    red_fill = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")  # Красный цвет
    
    for cell in sheet['G'][1:]:  # Пропускаем заголовок (начинаем с 1 строки)
        if cell.value == 1:
            cell.fill = green_fill
        elif cell.value == 0:
            cell.fill = red_fill

# Основная функция для создания Excel файла с улучшенной стилизацией
def create_excel_with_data(all_users, true_users, false_users, file_name="output.xlsx"):
    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        sheet_1 = workbook.create_sheet("Sheet_1")
        sheet_2 = workbook.create_sheet("Sheet_2")
        sheet_3 = workbook.create_sheet("Sheet_3")

        sheet_1.title = "All users"
        sheet_2.title = "True users"
        sheet_3.title = "False users"

        headers = ["id", "Telegram ID", "Name", "School", "City", "Number", "Payment", "Language"]
        sheet_1.append(headers)
        sheet_2.append(headers)
        sheet_3.append(headers)

        # Добавляем данные
        for row in all_users:
            sheet_1.append(row)

        for row in true_users:
            sheet_2.append(row)

        for row in false_users:
            sheet_3.append(row)
        
        columns_to_style = ["A", "B", "C", "D", "E", "F", "G", "H"]

        # Стилизация заголовков и данных
        style_header_cells(sheet_1, columns_to_style)
        style_header_cells(sheet_2, columns_to_style)
        style_header_cells(sheet_3, columns_to_style)

        style_body_cells(sheet_1, columns_to_style)
        style_body_cells(sheet_2, columns_to_style)
        style_body_cells(sheet_3, columns_to_style)

        # Раскрашиваем колонку 'G' в зависимости от значений в sheet_1
        style_column_g(sheet_1)

        # Сохраняем файл
        workbook.save(file_name)
        print(f"Файл '{file_name}' успешно создан.")
        return file_name
    except Exception as e:
        print(f"Ошибка при создании Excel файла: {e}")
        return None



# # Пример данных
# user_data = [
#     [123456789, "John Doe", "School 1", "City A", "+1234567890", "Paid", "English"],
#     [987654321, "Jane Smith", "School 2", "City B", "+0987654321", "Unpaid", "Spanish"]
# ]

# create_excel_with_data(user_data, "user_data.xlsx")


class chat_type_filter(Filter):
    def __init__(self,chat_types:list[str]) -> None:
        self.chat_types = chat_types

    async def __call__(self,message:Message) -> bool:
        return message.chat.type in self.chat_types


class MediaFilter(BaseFilter):
    async def __call__(self, message: Message) -> bool:
        file = message.content_type in [ContentType.PHOTO, ContentType.DOCUMENT]
        return file if file else 'False'